"""
检查当前 manifest 状态，统计各状态文件数。
"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"d:\programmtools\tools\ragsystem\data")
MANIFEST = DATA_DIR / "manifest.xlsx"

wb = openpyxl.load_workbook(str(MANIFEST), data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]

# 找到关键列
fname_col = headers.index("文件名称") if "文件名称" in headers else None
status_col = headers.index("status") if "status" in headers else None
parse_col = headers.index("parse") if "parse" in headers else None
dify_col = headers.index("dify_doc_id") if "dify_doc_id" in headers else None

print(f"关键列索引: 文件名称={fname_col}, status={status_col}, parse={parse_col}, dify_doc_id={dify_col}")

# 统计各状态
stats = {
    "total": 0,
    "pending": 0,
    "parsing": 0,
    "parsed": 0,
    "chunked": 0,
    "done": 0,
    "error": 0,
    "has_dify_id": 0,
    "no_dify_id": 0,
}

pending_files = []
error_files = []

for row in ws.iter_rows(min_row=2, values_only=True):
    fn = row[fname_col] if fname_col is not None else None
    if not fn or not str(fn).strip():
        continue
    
    stats["total"] += 1
    status = str(row[status_col]).strip() if status_col is not None and row[status_col] else ""
    parse_status = str(row[parse_col]).strip() if parse_col is not None and row[parse_col] else ""
    dify_id = str(row[dify_col]).strip() if dify_col is not None and row[dify_col] else ""
    
    if dify_id and dify_id != "None":
        stats["has_dify_id"] += 1
    else:
        stats["no_dify_id"] += 1
    
    if status == "done":
        stats["done"] += 1
    elif status == "error":
        stats["error"] += 1
        error_files.append(fn)
    elif status == "chunked":
        stats["chunked"] += 1
    elif status == "parsed" or parse_status == "done":
        stats["parsed"] += 1
    elif status == "parsing":
        stats["parsing"] += 1
    else:
        stats["pending"] += 1
        pending_files.append(fn)

print(f"\n=== Manifest 状态统计 ===")
print(f"总计: {stats['total']}")
print(f"  done (已上传 Dify): {stats['done']}")
print(f"  chunked (已切分待上传): {stats['chunked']}")
print(f"  parsed (已解析待切分): {stats['parsed']}")
print(f"  parsing (解析中): {stats['parsing']}")
print(f"  pending (待处理): {stats['pending']}")
print(f"  error (错误): {stats['error']}")
print(f"\n有 Dify ID: {stats['has_dify_id']}")
print(f"无 Dify ID: {stats['no_dify_id']}")

if pending_files:
    print(f"\n=== 待处理文件 ({len(pending_files)}) ===")
    for fn in pending_files[:20]:
        print(f"  {fn}")
    if len(pending_files) > 20:
        print(f"  ... 还有 {len(pending_files) - 20} 个")

if error_files:
    print(f"\n=== 错误文件 ({len(error_files)}) ===")
    for fn in error_files:
        print(f"  {fn}")

wb.close()
