"""修复未匹配的文件名映射。"""
import re
from pathlib import Path
from openpyxl import load_workbook

DATA = Path(r"d:\programmtools\tools\ragsystem\data")
INPUT = DATA / "input"
MANIFEST = DATA / "manifest.xlsx"

# 手动映射：manifest stem → input 实际文件名
MANUAL_MAP = {
    "A Multifaceted Approach to Education,Observation and Feedback in a Successful Hand Hygiene Campaign":
        "A_Multifaceted_Approach_to_Education_Observation_and_Feedback_in_a_Successful_Hand_Hygiene_Campaign_(2).pdf",
    "手卫生技术参考手册（WHO）": "手卫生技术参考手册_WHO.pdf",
    "GBT20810-2018 卫生纸（含卫生纸原纸）": "GBT20810-2018 卫生纸(含卫生纸原纸).pdf",
    "2013_PHAC 医疗保健环境中的手部卫生实践": "2013_PHAC_Hand Hygiene-EN医疗保健环境中的手部卫生实践 .pdf",
    "《清洁的手，呵护健康（2015-2018 年）》": "《清洁的手，呵护健康（ 20152015 2015-20182018 2018年）》.pdf",
    "GBT20808-2022纸巾": "GBT20808-2022纸巾.pdf",
    "GBT24455-2022 擦手纸": "GBT24455-2022擦手纸.pdf",
    "医护環境內保持手部衞生及使用手套的建議": "醫護環境內保持手部衞生及使用手套的建議(二零一七年十一月)(只備英文版)recommendations_on_hand_hygiene_and_use_of_gloves_in_health_care_settings.pdf",
    "医护人员手卫生规范_解读_李六亿": "医务人员手卫生规范_解读_李六亿.pdf",
}

wb = load_workbook(str(MANIFEST))
ws = wb.active
headers = [c.value for c in ws[1]]
fname_col = headers.index("文件名称")

fixed = 0
for row in ws.iter_rows(min_row=2):
    fname_cell = row[fname_col]
    fname = str(fname_cell.value or "").strip()
    if not fname:
        continue
    
    # 检查文件是否存在
    found = False
    for ext in ["", ".pdf", ".docx", ".doc"]:
        if (INPUT / (fname + ext)).exists():
            found = True
            break
    
    if found:
        continue
    
    # 尝试手动映射
    actual = MANUAL_MAP.get(fname)
    if actual and (INPUT / actual).exists():
        fname_cell.value = actual
        fixed += 1
        print(f"  修复: {fname} → {actual}")
        continue
    
    # 尝试更宽松的匹配
    fname_norm = re.sub(r'[\s\-_—（）()\[\]【】《》]', '', fname).lower()
    for f in INPUT.iterdir():
        if f.name == ".gitkeep":
            continue
        f_norm = re.sub(r'[\s\-_—（）()\[\]【】《》]', '', f.stem).lower()
        if fname_norm and f_norm and (fname_norm in f_norm or f_norm in fname_norm):
            fname_cell.value = f.name
            fixed += 1
            print(f"  模糊匹配: {fname} → {f.name}")
            found = True
            break

wb.save(str(MANIFEST))
wb.close()
print(f"\n共修复 {fixed} 个文件名")
