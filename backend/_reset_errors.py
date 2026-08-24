"""重置错误文件的 parse 列，让它们可以被重新解析。"""
import openpyxl
from pathlib import Path

MANIFEST = Path(r"d:\programmtools\tools\ragsystem\data\manifest.xlsx")
wb = openpyxl.load_workbook(str(MANIFEST))
ws = wb.active

headers = [c.value for c in ws[1]]
print(f"Headers: {headers[:15]}...")

# 找到相关列
col_map = {}
for i, h in enumerate(headers):
    if h:
        col_map[h] = i + 1

print(f"Columns: {list(col_map.keys())}")

parse_col = col_map.get("parse")
status_col = col_map.get("status")
fname_col = col_map.get("文件名称")

print(f"parse_col={parse_col}, status_col={status_col}, fname_col={fname_col}")

if not parse_col or not status_col:
    print("ERROR: 找不到 parse 或 status 列")
    exit(1)

# 重置 error 状态的文件
reset_count = 0
for row in ws.iter_rows(min_row=2):
    status_val = row[status_col - 1].value
    if status_val == "error":
        fname = row[fname_col - 1].value if fname_col else "?"
        row[parse_col - 1].value = ""
        row[status_col - 1].value = "pending"
        # 也清空 chunks 和 dify 列
        for h in ["chunks", "dify_doc_id", "dify_status"]:
            c = col_map.get(h)
            if c:
                row[c - 1].value = ""
        reset_count += 1
        print(f"  重置: {fname}")

wb.save(str(MANIFEST))
print(f"\n共重置 {reset_count} 个文件")
