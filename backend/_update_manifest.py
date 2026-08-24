"""
更新 manifest.xlsx，添加 pending/ 中新增的文件（从手卫生文件夹复制过来的）。
"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"d:\programmtools\tools\ragsystem\data")
PENDING = DATA_DIR / "pending"
MANIFEST = DATA_DIR / "manifest.xlsx"

# 1) 读取当前 manifest
wb = openpyxl.load_workbook(str(MANIFEST), data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
print(f"当前 manifest headers: {headers}")
print(f"当前 manifest 行数: {ws.max_row - 1}")

# 获取文件名称列
fname_col = headers.index("文件名称") if "文件名称" in headers else None
if fname_col is None:
    print("ERROR: 找不到 '文件名称' 列")
    sys.exit(1)

# 获取当前 manifest 中的文件名
manifest_files = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    fn = row[fname_col]
    if fn and str(fn).strip():
        manifest_files.add(str(fn).strip())

print(f"当前 manifest 文件数: {len(manifest_files)}")

# 2) 列出 pending/ 下所有文件
pending_files = []
for f in sorted(PENDING.iterdir()):
    if f.is_file() and f.name != ".gitkeep":
        pending_files.append(f)

print(f"pending/ 文件数: {len(pending_files)}")

# 3) 找出 pending/ 中不在 manifest 中的文件
missing = []
for f in pending_files:
    if f.name not in manifest_files:
        missing.append(f)

print(f"\n不在 manifest 中的文件: {len(missing)} 个")
for f in missing:
    print(f"  {f.name}")

wb.close()

# 4) 添加缺失文件到 manifest
if missing:
    wb = openpyxl.load_workbook(str(MANIFEST))
    ws = wb.active
    
    # 获取当前最大序号
    max_seq = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = row[0]
        if seq and isinstance(seq, (int, float)):
            max_seq = max(max_seq, int(seq))
    
    print(f"\n当前最大序号: {max_seq}")
    
    # 添加缺失文件
    for f in missing:
        max_seq += 1
        new_row = [max_seq, f.name, "", "", "", "", "", "", "", "", "", ""]
        ws.append(new_row)
        print(f"  Added: {f.name} (seq={max_seq})")
    
    wb.save(str(MANIFEST))
    wb.close()
    print(f"\n已更新 manifest.xlsx，新增 {len(missing)} 条记录")
else:
    print("\n没有需要添加的文件")

# 5) 显示最终统计
wb = openpyxl.load_workbook(str(MANIFEST), data_only=True)
ws = wb.active
final_count = ws.max_row - 1
print(f"\n最终 manifest 文件数: {final_count}")
wb.close()
