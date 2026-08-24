"""将元数据格式的 manifest.xlsx 转换为流水线格式，然后运行全流程。"""
import re
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook

DATA = Path(r"d:\programmtools\tools\ragsystem\data")
INPUT = DATA / "input"
MANIFEST = DATA / "manifest.xlsx"
BACKUP = DATA / "manifest_metadata_backup.xlsx"

# 1) 读取当前 manifest（元数据格式）
wb_old = load_workbook(str(MANIFEST), data_only=True)
ws_old = wb_old.active

meta_rows = []
for row in ws_old.iter_rows(min_row=3, values_only=True):
    stem = row[0]
    if stem and str(stem).strip():
        meta_rows.append({
            "stem": str(stem).strip(),
            "doc_type_primary": str(row[1] or "").strip(),
            "doc_type_secondary": str(row[2] or "").strip(),
            "topic_primary": str(row[3] or "").strip(),
            "topic_secondary": str(row[4] or "").strip(),
            "core_summary": str(row[5] or "").strip(),
            "entity_label": str(row[6] or "").strip(),
            "attribute_label": str(row[7] or "").strip(),
            "applicable_scenarios": str(row[8] or "").strip(),
            "effective_date": str(row[9] or "").strip(),
            "priority": row[10] if row[10] is not None else "",
            "status": str(row[11] or "").strip(),
        })
wb_old.close()
print(f"读取到 {len(meta_rows)} 条元数据记录")

# 2) 备份旧 manifest
shutil.copy2(str(MANIFEST), str(BACKUP))
print(f"已备份到: {BACKUP}")

# 3) 匹配 input/ 中的实际文件
def find_input_file(stem: str) -> str | None:
    """根据 stem 在 input/ 中查找匹配的文件（模糊匹配）。"""
    # 标准化 stem（去空格、特殊字符）
    stem_norm = re.sub(r'[\s\-_—]', '', stem).lower()
    
    for f in INPUT.iterdir():
        if f.name == ".gitkeep":
            continue
        file_stem = f.stem
        file_stem_norm = re.sub(r'[\s\-_—]', '', file_stem).lower()
        
        # 精确匹配（标准化后）
        if file_stem_norm == stem_norm:
            return f.name
        # stem 包含匹配
        if stem_norm in file_stem_norm or file_stem_norm in stem_norm:
            return f.name
    
    return None

matched = []
unmatched = []
for meta in meta_rows:
    actual_name = find_input_file(meta["stem"])
    if actual_name:
        matched.append({**meta, "filename": actual_name})
    else:
        unmatched.append(meta["stem"])

print(f"\n匹配成功: {len(matched)} 个文件")
print(f"未匹配: {len(unmatched)} 个")
if unmatched:
    print("未匹配的文件:")
    for s in unmatched:
        print(f"  - {s}")

# 4) 创建新的 manifest.xlsx（流水线格式）
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "Sheet1"

# 标准列 + 元数据列
headers = [
    "序号", "文件名称", "一级分类", "二级分类", "关键词标签",
    "适用科室", "生效日期", "导入情况", "处理情况", "校对", "处理说明",
    # 元数据列（追加在后面）
    "doc_type_primary", "doc_type_secondary", "topic_primary", "topic_secondary",
    "core_summary", "entity_label", "attribute_label", "applicable_scenarios",
    "priority", "dify_meta_status",
]
ws_new.append(headers)

for i, m in enumerate(matched):
    ws_new.append([
        i + 1,                              # 序号
        m["filename"],                       # 文件名称
        m["doc_type_primary"],               # 一级分类
        m["doc_type_secondary"],             # 二级分类
        m["entity_label"],                   # 关键词标签
        m["applicable_scenarios"],           # 适用科室
        m["effective_date"] if m["effective_date"] != "无" else "",  # 生效日期
        "",                                  # 导入情况
        "",                                  # 处理情况
        "",                                  # 校对
        "",                                  # 处理说明
        # 元数据列
        m["doc_type_primary"],
        m["doc_type_secondary"],
        m["topic_primary"],
        m["topic_secondary"],
        m["core_summary"],
        m["entity_label"],
        m["attribute_label"],
        m["applicable_scenarios"],
        m["priority"],
        "",  # dify_meta_status
    ])

wb_new.save(str(MANIFEST))
wb_new.close()
print(f"\n新 manifest.xlsx 已创建: {MANIFEST}")
print(f"  - 共 {len(matched)} 行数据")
print(f"  - 列: {headers}")

# 5) 同时创建 doc_metadata.xlsx（供 dify 元数据写入用）
wb_meta = Workbook()
ws_meta = wb_meta.active
ws_meta.title = "Sheet1"

# 第1行：英文字段名（第1格为空）
ws_meta.append([None, "doc_type_primary", "doc_type_secondary", "topic_primary",
                "topic_secondary", "core_summary", "entity_label", "attribute_label",
                "applicable_scenarios", "effective_date", "priority", "status"])
# 第2行：中文列名
ws_meta.append(["文件名", "类型-一级", "类型-二级", "主题-一级", "主题-二级",
                "核心内容摘要", "实体标签", "属性标签", "适用科室", "生效日期", "", ""])

for m in matched:
    # stem 用实际文件的 stem（不含后缀）
    file_stem = Path(m["filename"]).stem
    priority_val = m["priority"]
    if priority_val != "" and priority_val is not None:
        try:
            priority_val = float(priority_val)
        except (TypeError, ValueError):
            priority_val = ""
    ws_meta.append([
        file_stem,
        m["doc_type_primary"],
        m["doc_type_secondary"],
        m["topic_primary"],
        m["topic_secondary"],
        m["core_summary"],
        m["entity_label"],
        m["attribute_label"],
        m["applicable_scenarios"],
        m["effective_date"] if m["effective_date"] != "无" else "",
        priority_val,
        m["status"],
    ])

meta_path = DATA / "doc_metadata.xlsx"
wb_meta.save(str(meta_path))
wb_meta.close()
print(f"\ndoc_metadata.xlsx 已创建: {meta_path}")
print(f"  - 共 {len(matched)} 条记录")
