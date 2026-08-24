"""
检查 manifest 中的重复条目并清理。
"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"d:\programmtools\tools\ragsystem\data")
MANIFEST = DATA_DIR / "manifest.xlsx"

wb = openpyxl.load_workbook(str(MANIFEST))
ws = wb.active
headers = [c.value for c in ws[1]]

fname_col = headers.index("文件名称") if "文件名称" in headers else None
status_col = headers.index("status") if "status" in headers else None
dify_col = headers.index("dify_doc_id") if "dify_doc_id" in headers else None

# 收集所有条目
entries = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    fn = row[fname_col] if fname_col is not None else None
    if not fn or not str(fn).strip():
        continue
    status = str(row[status_col]).strip() if status_col is not None and row[status_col] else ""
    dify_id = str(row[dify_col]).strip() if dify_col is not None and row[dify_col] else ""
    entries.append({
        "row_idx": row_idx,
        "filename": str(fn).strip(),
        "status": status,
        "dify_id": dify_id,
    })

print(f"总条目: {len(entries)}")

# 检查重复文件名
from collections import Counter
fname_counts = Counter(e["filename"] for e in entries)
duplicates = {fn: count for fn, count in fname_counts.items() if count > 1}

if duplicates:
    print(f"\n=== 重复文件名 ({len(duplicates)}) ===")
    for fn, count in duplicates.items():
        print(f"  {fn}: {count} 次")
        for e in entries:
            if e["filename"] == fn:
                print(f"    Row {e['row_idx']}: status={e['status']}, dify_id={e['dify_id']}")
else:
    print("\n没有重复文件名")

# 检查 pending 文件夹中的实际文件
PENDING = DATA_DIR / "pending"
pending_files = []
for f in sorted(PENDING.iterdir()):
    if f.is_file() and f.name != ".gitkeep":
        pending_files.append(f.name)

print(f"\npending/ 文件数: {len(pending_files)}")

# 检查 manifest 中哪些条目对应的文件不在 pending/ 中
pending_set = set(pending_files)
not_in_pending = []
for e in entries:
    if e["filename"] not in pending_set:
        not_in_pending.append(e)

if not_in_pending:
    print(f"\n=== manifest 条目但文件不在 pending/ 中 ({len(not_in_pending)}) ===")
    for e in not_in_pending:
        print(f"  Row {e['row_idx']}: {e['filename']} (status={e['status']})")

wb.close()
