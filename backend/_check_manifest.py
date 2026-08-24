"""Check manifest files against input/."""
import openpyxl
from pathlib import Path

INPUT = Path(r"d:\programmtools\tools\ragsystem\data\input")
wb = openpyxl.load_workbook(r"d:\programmtools\tools\ragsystem\data\manifest.xlsx", data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
fname_col = headers.index("文件名称")

print("Checking manifest files against input/:")
exist = 0
missing = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    fname = str(row[fname_col] or "").strip()
    if not fname:
        continue
    found = (INPUT / fname).exists()
    if found:
        exist += 1
    else:
        missing += 1
        print(f"  MISSING: {fname}")

print(f"\nExist: {exist}, Missing: {missing}")
input_files = [f for f in INPUT.iterdir() if f.name != ".gitkeep"]
print(f"Total in input/: {len(input_files)}")
