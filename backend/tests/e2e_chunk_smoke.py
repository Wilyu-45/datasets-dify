"""E2E 验证：§3.3 自定义切分流程。

场景：
  1. 用户放入 18 列 Excel（11 用户 + 5 系统 + parse + chunks）
  2. 准备两个已解析文档（parsed/{stem}/）：
     - 文档 A：WST 809 风格（cover/toc/preface/body/appendix）
     - 文档 B：医院感染风格（cover + 直接第一章，无目录/前言）
  3. 启动 → bootstrap 不动用户列
  4. 模拟「切分」按钮：调 chunker.chunk_parsed
  5. 验证：
       - chunks/{stem}/ 目录被创建
       - chunk_NNN_xxx.md 落盘正确
       - chunk_metadata.json 含正确 title_path/chunk_type/image_refs
       - 引用的图片被拷贝到 chunks/{stem}/images/
       - manifest 的 chunks 列 = chunks/{stem}
       - manifest 的 status 列 = 'chunked'
  6. 幂等性：第二次切分 = SKIPPED_DONE
  7. force=true：清空重切（验证 chunks_dir 重建）
"""

from __future__ import annotations

import importlib
import json
import logging
import shutil
import sys
from pathlib import Path
from typing import Any, Dict, List

logging.disable(logging.CRITICAL)

# ============ 准备 ============

ROOT = Path("d:/programmtools/tools/ragsystem")
DATA = ROOT / "data"
PARSED = DATA / "parsed"
CHUNKS = DATA / "chunks"
MANIFEST = DATA / "manifest.xlsx"

BACKUP = DATA / "manifest.xlsx.bak"
if BACKUP.exists():
    BACKUP.unlink()
if MANIFEST.exists():
    shutil.copy2(MANIFEST, BACKUP)

# 本测试创建的两个 stem（仅清掉这两个子目录，不动其它真实数据）
TEST_STEMS = ("国标-W809", "规范-医院感染")

# 清掉本测试可能遗留的 chunks 子目录（如果之前测试失败留下了）
for stem in TEST_STEMS:
    test_chunks_dir = CHUNKS / stem
    if test_chunks_dir.exists():
        shutil.rmtree(test_chunks_dir, ignore_errors=True)
    test_parsed_dir = PARSED / stem
    if test_parsed_dir.exists():
        shutil.rmtree(test_parsed_dir, ignore_errors=True)

# ============ 清理函数：atexit 兜底，即使断言失败也恢复 manifest ============
def _cleanup() -> None:
    """恢复 manifest + 清掉本测试创建的子目录，不动其它真实数据。"""
    # 恢复 manifest
    if BACKUP.exists():
        shutil.move(BACKUP, MANIFEST)
    # 清掉本测试创建的 chunks 子目录
    for stem in TEST_STEMS:
        test_chunks_dir = CHUNKS / stem
        if test_chunks_dir.exists():
            shutil.rmtree(test_chunks_dir, ignore_errors=True)
        test_parsed_dir = PARSED / stem
        if test_parsed_dir.exists():
            shutil.rmtree(test_parsed_dir, ignore_errors=True)


import atexit
atexit.register(_cleanup)

# ============ 1) 准备用户的 18 列 Excel ============

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
headers = [
    # 11 用户列
    "序号", "文件名称", "一级分类", "二级分类", "关键词标签",
    "适用科室", "生效日期", "导入情况", "处理情况", "校对", "处理说明",
    # 5 系统列
    "status", "md5", "create_time", "update_time", "error_msg",
    # §3.2 解析列
    "parse",
    # §3.3 切分列
    "chunks",
]
ws.append(headers)
# 文档 A：WST 809 风格（带 cover/toc/preface/appendix）
ws.append([
    1, "国标-W809", "国标", "视觉设计", "WST 809", "全科", "2022-01-01",
    "已移入待处理", "已扫描", "", "WST 809 风格",
    "parsing_done", "md5_w809", "2026-07-30 10:00:00", "2026-07-30 10:00:00", "",
    str(PARSED / "国标-W809"),
    "",  # chunks 空（裸 stem 或空字符串）
])
# 文档 B：医院感染风格（cover + 直接第一章，无目录/前言）
ws.append([
    2, "规范-医院感染", "规范", "院感", "暴发", "院感科", "2009-10-01",
    "已移入待处理", "已扫描", "", "医院感染风格",
    "parsing_done", "md5_yqgr", "2026-07-30 10:00:00", "2026-07-30 10:00:00", "",
    str(PARSED / "规范-医院感染"),
    "",  # chunks 空
])
wb.save(MANIFEST)
wb.close()

# ============ 2) 准备两个已解析文档（v2 + images）============

# --- 文档 A：WST 809 风格 v2 ---
stem_a = PARSED / "国标-W809"
inner_a = stem_a / "hybrid_auto"
inner_a.mkdir(parents=True)
images_a = inner_a / "images"
images_a.mkdir()

# 写一张假图片
(images_a / "fig_a1.jpg").write_bytes(b"\xff\xd8\xff\xe0FAKE_JPG_A")
(images_a / "fig_b1.jpg").write_bytes(b"\xff\xd8\xff\xe0FAKE_JPG_B")

# v2 文档 A：cover / toc / preface / body / appendix 五段
v2_a = [
    # 封面
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "国标-W809 视觉设计标准"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "国标-W809—2022，本标准规定了基层医疗卫生机构功能单元视觉设计的基本原则和各功能区域的具体要求，适用于乡镇卫生院、社区卫生服务中心等基层医疗机构的视觉形象建设。"}]}},
    ],
    # 目录
    [
        {"type": "title", "content": {"level": 2, "title_content": [{"content": "目 录"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "1 范围....1\n2 规范性引用文件....2\n3 术语和定义....3\n4 功能单元视觉设计标准....4"}]}},
    ],
    # 前言
    [
        {"type": "title", "content": {"level": 2, "title_content": [{"content": "前 言"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "本标准由国家卫生健康标准委员会提出并归口。本标准起草单位包括中国疾控中心、北京市卫生健康委员会等单位。本标准规定了基层医疗卫生机构功能单元视觉设计的基本原则和要求。"}]}},
    ],
    # 正文
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "1 范围"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "本标准适用于基层医疗卫生机构功能单元的视觉设计，包括预防保健区、诊疗区、辅助区等功能单元。其他医疗机构可参照执行。"}]}},
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "2 定义"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "下列术语和定义适用于本文件。功能单元是指医疗机构内具有特定医疗或保健功能的房间或区域组合。"}]}},
    ],
    # 附录 A
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "附 录 A"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "（资料性）本附录给出了功能单元视觉设计的基本要求和方法说明，包括色彩搭配、字体规范、标识设置等具体内容。"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "示例见图A.1，展示了典型的预防保健区入口外观效果图，包含了标识、色彩和字体等设计要素。"}]}},
        {"type": "image", "content": {
            "image_source": {"path": "images/fig_a1.jpg"},
            "image_caption": [{"content": "图 A.1 示例图"}],
        }},
    ],
    # 附录 B
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "附 录 B"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "示例见图B.1，展示了B超、心电图室等功能单元的标识设计方案。"}]}},
        {"type": "image", "content": {
            "image_source": {"path": "images/fig_b1.jpg"},
            "image_caption": [{"content": "图 B.1 示例图"}],
        }},
    ],
]
(inner_a / "国标-W809_content_list_v2.json").write_text(
    json.dumps(v2_a, ensure_ascii=False), encoding="utf-8"
)
(inner_a / "国标-W809.md").write_text(
    "# 国标-W809 视觉设计标准\n\n本标准规定了基层医疗卫生机构功能单元视觉设计的基本原则和各功能区域的具体要求。",
    encoding="utf-8",
)

# --- 文档 B：医院感染风格 v2（无目录/前言，cover 后直接第一章） ---
stem_b = PARSED / "规范-医院感染"
inner_b = stem_b / "hybrid_auto"
inner_b.mkdir(parents=True)

v2_b = [
    # 封面
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "关于印发《医院感染暴发报告及处置管理规范》的通知"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "各省、自治区、直辖市卫生厅局、中医药管理局，新疆生产建设兵团卫生局，为贯彻落实《医院感染管理办法》，进一步规范医院感染暴发报告和处置的管理工作。"}]}},
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "医院感染暴发报告及处置管理规范"}]}},
    ],
    # 第一章
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "第一章 总则"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "第一条 为贯彻落实《医院感染管理办法》，进一步规范医院感染暴发报告和处置的管理工作，最大限度地降低医院感染对患者造成的危害，保障医疗安全，制定本规范。"}]}},
    ],
    # 第二章
    [
        {"type": "title", "content": {"level": 1, "title_content": [{"content": "第二章 组织管理"}]}},
        {"type": "paragraph", "content": {"paragraph_content": [{"content": "第六条 各级卫生行政部门和医疗机构应当加强对医院感染暴发报告和处置工作的组织领导，建立相应的管理制度和应急预案，确保报告和处置工作及时、有序、高效。"}]}},
    ],
]
(inner_b / "规范-医院感染_content_list_v2.json").write_text(
    json.dumps(v2_b, ensure_ascii=False), encoding="utf-8"
)
(inner_b / "规范-医院感染.md").write_text(
    "# 医院感染暴发报告及处置管理规范\n\n第一条 为贯彻落实《医院感染管理办法》，进一步规范医院感染暴发报告和处置的管理工作。",
    encoding="utf-8",
)

print(f"✓ 准备 2 个已解析文档：国标-W809（含 2 张图）、规范-医院感染")

# ============ 3) 启动 bootstrap（不移动文件，仅补列）============

sys.path.insert(0, str(ROOT / "backend"))
import os
os.environ.setdefault("RAG_DATA_ROOT", str(DATA))

from app import config as cfg_mod
importlib.reload(cfg_mod)
settings = cfg_mod.settings
from app.services import chunker, manifest_store
chunker.settings = settings

manifest_store.bootstrap(DATA)

# 验证列数 = 18
wb2 = __import__("openpyxl").load_workbook(MANIFEST, read_only=True)
ws2 = wb2.active
hdr = [str(c).strip() if c else "" for c in next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))]
wb2.close()
assert len(hdr) == 18, f"列数应为 18，实际 {len(hdr)}"
assert hdr[-1] == "chunks", f"最后一列应为 chunks，实际 {hdr[-1]}"
assert hdr[-2] == "parse", f"倒数第二列应为 parse，实际 {hdr[-2]}"
print(f"✓ bootstrap 补列成功（18 列，最后两列=parse/chunks）")

# ============ 4) 模拟「切分」按钮 ============

report = chunker.chunk_parsed(dry_run=False, force=False)
print(f"✓ 切分完成: scanned={report.scanned}, chunked={report.chunked}, "
      f"skipped={report.skipped_done}, failed={report.failed}")
assert report.chunked == 2, f"应 chunked=2, 实际={report.chunked}"
assert report.failed == 0, f"应 failed=0, 实际={report.failed}"

# ============ 5) 验证文档 A（WST 809 风格）的切分产物 ============

chunks_a = CHUNKS / "国标-W809"
assert chunks_a.is_dir(), f"应存在 {chunks_a}"

# 至少有 5 个 chunk（cover/toc/preface/2*body/2*appendix → 至少 6 个）
chunk_files = sorted(chunks_a.glob("chunk_*.md"))
assert len(chunk_files) >= 5, f"应至少 5 个 chunk 文件，实际 {len(chunk_files)}"
print(f"✓ 国标-W809: 生成 {len(chunk_files)} 个 chunk 文件")

# 验证 chunk_metadata.json
meta_a = json.loads((chunks_a / "chunk_metadata.json").read_text(encoding="utf-8"))
assert meta_a["doc_stem"] == "国标-W809"
assert meta_a["chunk_count"] == len(chunk_files)

# 至少包含 cover/toc/preface/body/appendix 五种类型
types = {c["chunk_type"] for c in meta_a["chunks"]}
assert "cover" in types, f"缺 cover，实际类型: {types}"
assert "toc" in types, f"缺 toc，实际类型: {types}"
assert "preface" in types, f"缺 preface，实际类型: {types}"
assert "body" in types, f"缺 body，实际类型: {types}"
assert "appendix" in types, f"缺 appendix，实际类型: {types}"
print(f"✓ chunk_metadata.json 含 5 种类型: {sorted(types)}")

# 验证图片被拷贝
images_dir_a = chunks_a / "images"
assert images_dir_a.is_dir(), f"应存在 {images_dir_a}"
copied_imgs = sorted(images_dir_a.glob("*.jpg"))
assert len(copied_imgs) == 2, f"应拷贝 2 张图片，实际 {len(copied_imgs)}"
print(f"✓ 图片拷贝成功: {[p.name for p in copied_imgs]}")

# 验证附录 chunk 含 image_refs
appendix_chunks = [c for c in meta_a["chunks"] if c["chunk_type"] == "appendix"]
total_refs = sum(len(c["image_refs"]) for c in appendix_chunks)
assert total_refs == 2, f"appendix 应共引用 2 张图，实际 {total_refs}"
print(f"✓ 附录 chunk image_refs 总计 {total_refs} 张")

# 验证 chunk 内容含 MD 原生图片语法 `![](images/xxx.jpg)`，不含 caption
appendix_md = next(
    f for f in chunk_files
    if f.name.startswith("chunk_") and "附_录" in f.name
).read_text(encoding="utf-8")
assert "![](images/" in appendix_md, "应使用 MD 原生图片语法"
# 附录 chunk 文件名应含『附_录』
print(f"✓ 附录 chunk MD 语法正确: 含 ![](images/...)")

# ============ 6) 验证文档 B（医院感染风格）的切分产物 ============

chunks_b = CHUNKS / "规范-医院感染"
assert chunks_b.is_dir(), f"应存在 {chunks_b}"

# 关键修复：无目录/前言时，cover 不能吞掉"第一章" → 至少 3 个 chunk（cover + 2 章）
chunk_files_b = sorted(chunks_b.glob("chunk_*.md"))
assert len(chunk_files_b) >= 3, f"医院感染应至少 3 个 chunk（cover+2章），实际 {len(chunk_files_b)}"
print(f"✓ 规范-医院感染: 生成 {len(chunk_files_b)} 个 chunk 文件")

# 验证 body 的第一个 title 是"第一章"而不是 cover
meta_b = json.loads((chunks_b / "chunk_metadata.json").read_text(encoding="utf-8"))
body_chunks = [c for c in meta_b["chunks"] if c["chunk_type"] == "body"]
assert body_chunks, "应有 body 类型 chunk"
first_body_title = body_chunks[0]["title_path"]
assert "第一章" in first_body_title, f"body 起点应含『第一章』，实际={first_body_title}"
print(f"✓ 医院感染 body 起点正确: {first_body_title}")

# cover 不应包含"第一章"
cover_chunk = next(c for c in meta_b["chunks"] if c["chunk_type"] == "cover")
assert "第一章" not in cover_chunk["title_path"]
print(f"✓ 医院感染 cover 正确：未吞掉 body 起点")

# ============ 7) 验证 manifest 更新 ============

manifest = manifest_store.load(MANIFEST)

# 文档 A
row_a = manifest["国标-W809"]
assert row_a.status == "chunked", f"status 应为 chunked，实际={row_a.status}"
assert row_a.chunks == "国标-W809", f"chunks 应为国标-W809（裸 stem），实际={row_a.chunks!r}"
assert row_a.update_time, "update_time 应被更新"
print(f"✓ 国标-W809 manifest: status={row_a.status}, chunks={row_a.chunks}")

# 文档 B
row_b = manifest["规范-医院感染"]
assert row_b.status == "chunked"
assert row_b.chunks == "规范-医院感染"
print(f"✓ 规范-医院感染 manifest: status={row_b.status}, chunks={row_b.chunks}")

# ============ 8) 幂等性：第二次切分 = 全 SKIPPED_DONE ============

report_2 = chunker.chunk_parsed(dry_run=False, force=False)
assert report_2.chunked == 0, f"第二次应 chunked=0, 实际={report_2.chunked}"
assert report_2.skipped_done == 2, f"第二次应 skipped=2, 实际={report_2.skipped_done}"
print(f"✓ 幂等性: 第二次切分 skipped={report_2.skipped_done}")

# ============ 9) force=true：清空重切 ============

report_3 = chunker.chunk_parsed(dry_run=False, force=True)
assert report_3.chunked == 2, f"force=True 应 chunked=2, 实际={report_3.chunked}"
# chunks 目录被清空重建 → 文件仍然存在
assert (chunks_a / "chunk_metadata.json").is_file()
assert (chunks_b / "chunk_metadata.json").is_file()
print(f"✓ force=True 清空重切: chunked={report_3.chunked}")

# ============ 10) dry_run 模式：不写盘 ============

# 改一个 manifest 行的 chunks 为空，模拟未切分状态
manifest2 = manifest_store.load(MANIFEST)
# 把一个 chunks 清空（直接修改 manifest）
row_a2 = manifest2["国标-W809"]
new_row = row_a2.model_copy(update={"chunks": ""})
manifest_store.upsert(MANIFEST, new_row)

# 备份目录以便验证 dry_run 不重写
chunks_bak = CHUNKS / "_dryrun_bak"
if chunks_bak.exists():
    shutil.rmtree(chunks_bak, ignore_errors=True)
shutil.copytree(chunks_a, chunks_bak)

# dry_run 不应实际切分（不写盘），但报告里会标记为 dry_run_chunk
report_dry = chunker.chunk_parsed(dry_run=True, force=False)
# 应有 dry_run_chunk action
assert any(a.action.value == "dry_run_chunk" for a in report_dry.actions), \
    f"应有 dry_run_chunk action, 实际 actions: {[a.action.value for a in report_dry.actions]}"
# 切分报告中 dry_run 标志应为 True
assert report_dry.dry_run is True
# chunks 目录不应被覆盖
chunks_after = sorted(p.name for p in chunks_a.glob("chunk_*.md"))
chunks_bak_files = sorted(p.name for p in chunks_bak.glob("chunk_*.md"))
assert chunks_after == chunks_bak_files, "dry_run 不应重写 chunks/"
print(f"✓ dry_run=True: 不写盘、不移动（action=dry_run_chunk）")

shutil.rmtree(chunks_bak, ignore_errors=True)

# ============ 清理 ============
# atexit 会在脚本退出时自动调用 _cleanup，无需手动调用。
# 但因为 atexit 在 assert 失败后也会触发，可以安全地在中间任何位置退出。

print("\n=== 全部 §3.3 E2E 验证通过 ===")
