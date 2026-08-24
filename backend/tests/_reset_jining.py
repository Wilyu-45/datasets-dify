"""Reset manifest + clean parsed for Jining PDF."""
import shutil
from pathlib import Path
from openpyxl import load_workbook

# 1. Clean parsed dir
parsed_dir = Path(r"D:\programmtools\tools\ragsystem\data\parsed\济宁市医疗卫生机构病死婴幼儿遗体处理暂行办法(1)")
if parsed_dir.exists():
    shutil.rmtree(parsed_dir)
    print(f"cleaned: {parsed_dir}")
else:
    print(f"not exists: {parsed_dir}")

# 2. Reset manifest parse column
manifest_path = Path(r"D:\programmtools\tools\ragsystem\data\manifest.xlsx")
wb = load_workbook(manifest_path)  # NOT read_only
ws = wb.active
headers = {cell.value: cell.column for cell in ws[1]}
parse_col = headers.get("parse")
status_col = headers.get("status")
print(f"parse_col={parse_col}, status_col={status_col}")

for row in ws.iter_rows(min_row=2):
    if row[headers["文件名称"]-1].value and "济宁" in str(row[headers["文件名称"]-1].value):
        old_parse = row[parse_col-1].value
        old_status = row[status_col-1].value
        row[parse_col-1].value = ""
        row[status_col-1].value = "pending"
        print(f"reset Jining: parse={old_parse!r} → '', status={old_status!r} → pending")
        break
else:
    print("Jining row not found")
wb.save(manifest_path)
print("manifest saved")
