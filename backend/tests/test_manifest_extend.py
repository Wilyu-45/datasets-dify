"""manifest_store 自动补列 + 灵活加载的测试。"""

from __future__ import annotations

import importlib
import logging
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

logging.disable(logging.CRITICAL)


@pytest.fixture(autouse=True)
def fresh_settings(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    test_data_root = tmp_path / "data"
    monkeypatch.setenv("RAG_DATA_ROOT", str(test_data_root))
    from app import config as cfg_mod
    importlib.reload(cfg_mod)
    settings = cfg_mod.settings
    from app.services import scanner
    scanner.settings = settings
    settings.ensure_dirs()
    yield settings


def _write_user_excel(
    path: Path,
    rows: list[dict],
    headers: list[str] | None = None,
) -> None:
    """模拟用户放入 data/ 的 Excel（11 列）。"""
    if headers is None:
        headers = [
            "序号", "文件名称", "一级分类", "二级分类", "关键词标签",
            "适用科室", "生效日期", "导入情况", "处理情况", "校对", "处理说明",
        ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _read_headers(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return [str(c).strip() if c is not None else "" for c in row]
    finally:
        wb.close()


# ============ ensure_columns 测试 ============


def test_bootstrap_extends_11_columns_to_20(fresh_settings):
    """用户放入 11 列 Excel → 启动 bootstrap → 自动补 9 列到末尾（5 系统 + parse + chunks + dify_doc_id + dify_status）。"""
    from app.services import manifest_store

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    _write_user_excel(
        user_path,
        rows=[
            {
                "序号": 1, "文件名称": "GBT 12130-2005.pdf",
                "一级分类": "国标", "二级分类": "医用氧舱",
                "关键词标签": "医用空气加压氧舱", "适用科室": "高压氧科",
                "生效日期": "2006/4/1", "导入情况": "", "处理情况": "",
                "校对": "", "处理说明": "扫描件、大文件",
            },
        ],
    )
    # 启动
    manifest_store.bootstrap(s.data_root)
    headers = _read_headers(user_path)
    # 11 列保留在前 11 位 + 5 个系统列 + 1 parse + 1 chunks + 2 dify = 20
    assert len(headers) == 20
    assert headers[:11] == [
        "序号", "文件名称", "一级分类", "二级分类", "关键词标签",
        "适用科室", "生效日期", "导入情况", "处理情况", "校对", "处理说明",
    ]
    assert headers[11:16] == ["status", "md5", "create_time", "update_time", "error_msg"]
    assert headers[16] == "parse"
    assert headers[17] == "chunks"
    assert headers[18] == "dify_doc_id"
    assert headers[19] == "dify_status"


def test_bootstrap_idempotent_on_20_columns(fresh_settings):
    """已经是 20 列 → bootstrap 不重复补列。"""
    from app.services import manifest_store

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    _write_user_excel(
        user_path,
        rows=[{"文件名称": "a.pdf"}],
    )
    manifest_store.bootstrap(s.data_root)
    headers_before = _read_headers(user_path)
    # 再 bootstrap
    changed, new_headers = manifest_store.ensure_columns(user_path)
    assert changed is False
    assert new_headers == headers_before
    assert _read_headers(user_path) == headers_before


def test_ensure_columns_partial(fresh_settings):
    """用户 Excel 缺部分列 → 只补缺失的，不动已存在的。"""
    from app.services import manifest_store

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    # 写一个只有 文件名称 和 导入情况 两列的 xlsx
    _write_user_excel(
        user_path,
        rows=[{"文件名称": "x.pdf", "导入情况": "已导入"}],
        headers=["文件名称", "导入情况"],
    )
    changed, new_headers = manifest_store.ensure_columns(user_path)
    assert changed is True
    # 原有的 2 列必须保留在前两位
    assert new_headers[:2] == ["文件名称", "导入情况"]
    # 缺失列追加在末尾：20 - 2 = 18 个
    assert len(new_headers) == 20
    # 检查处理类列在末尾
    assert new_headers[-1] == "dify_status"
    assert new_headers[-2] == "dify_doc_id"
    assert new_headers[-3] == "chunks"
    assert new_headers[-4] == "parse"
    assert "status" in new_headers
    assert "md5" in new_headers


def test_load_with_reordered_columns(fresh_settings):
    """列顺序打乱 → load 仍能正确识别 filename 字段。"""
    from app.services import manifest_store

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    # 反转顺序写
    headers = [
        "处理说明", "校对", "处理情况", "导入情况", "生效日期",
        "适用科室", "关键词标签", "二级分类", "一级分类", "文件名称", "序号",
    ]
    _write_user_excel(
        user_path,
        rows=[{"文件名称": "reordered.pdf", "序号": 99, "处理说明": "备注"}],
        headers=headers,
    )
    manifest = manifest_store.load(user_path)
    assert "reordered.pdf" in manifest
    row = manifest["reordered.pdf"]
    assert row.filename == "reordered.pdf"
    assert row.seq == 99
    assert row.process_note == "备注"


def test_load_preserves_user_data_after_extension(fresh_settings):
    """补列后用户原数据不变。"""
    from app.services import manifest_store

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    _write_user_excel(
        user_path,
        rows=[
            {
                "序号": 7,
                "文件名称": "TCAME 76.pdf",
                "一级分类": "团标",
                "二级分类": "医用氧舱",
                "关键词标签": "高原微压氧舱",
                "适用科室": "高压氧科",
                "生效日期": "2025/9/15",
                "导入情况": "",
                "处理情况": "",
                "校对": "",
                "处理说明": "新文件待处理",
            },
        ],
    )
    manifest_store.bootstrap(s.data_root)
    manifest = manifest_store.load(user_path)
    row = manifest["TCAME 76.pdf"]
    assert row.category_l1 == "团标"
    assert row.category_l2 == "医用氧舱"
    assert row.keywords == "高原微压氧舱"
    assert row.department == "高压氧科"
    assert row.effective_date == "2025/9/15"
    assert row.process_note == "新文件待处理"
    # 补的列都是 None
    assert row.status is None
    assert row.md5 is None
    assert row.create_time is None


# ============ scanner 与 manifest 协同测试 ============


def test_scan_updates_user_columns(fresh_settings):
    """扫描后 manifest 的『导入情况』『处理情况』『处理说明』被正确更新。"""
    from app.services import manifest_store, scanner

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    _write_user_excel(
        user_path,
        rows=[{
            "序号": 1, "文件名称": "new.pdf",
            "一级分类": "国标", "二级分类": "医用氧舱",
            "适用科室": "高压氧科", "处理说明": "原备注",
        }],
    )
    # 放一个文件到 input
    (s.input_dir / "new.pdf").write_bytes(b"hello world")
    report = scanner.scan_and_stage(dry_run=False)
    assert report.staged == 1
    manifest = manifest_store.load(user_path)
    row = manifest["new.pdf"]
    # 用户原列被系统更新
    assert row.import_status == "已移入待处理"
    assert row.process_status == "已扫描"
    # 处理说明被覆盖为 md5 摘要（系统行为）
    assert "md5=" in (row.process_note or "")
    # 原备注的「原备注」虽然丢了，但用户可在 Excel 里手动改回
    # 系统 5 列
    assert row.status == "pending"
    assert row.md5 and len(row.md5) == 32
    assert row.create_time
    assert row.update_time


def test_dry_run_sets_dry_run_marker_in_user_columns(fresh_settings):
    """dry_run=True 时不应写盘到 manifest，也不移动文件。"""
    from app.services import manifest_store, scanner

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    # 写一个 manifest 行（导入情况留空）
    _write_user_excel(user_path, rows=[{"文件名称": "dry.pdf"}])
    (s.input_dir / "dry.pdf").write_bytes(b"dry data")
    report = scanner.scan_and_stage(dry_run=True)
    assert report.dry_run is True
    # dry_run 不应写盘 → 导入情况仍为空
    manifest = manifest_store.load(user_path)
    row = manifest["dry.pdf"]
    assert row.import_status is None
    # 但文件不移动
    assert (s.input_dir / "dry.pdf").exists()
    # 报告里能看到这是 dry_run
    assert any(a.action.value == "dry_run" for a in report.actions)


def test_collision_rename_marks_in_user_columns(fresh_settings):
    """pending/ 同名但 md5 不同时 → 触发重命名，用户列写『已移入待处理(重命名)』。"""
    from app.services import manifest_store, scanner

    s = fresh_settings
    user_path = manifest_store.find_manifest_file(s.data_root)
    _write_user_excel(user_path, rows=[{"文件名称": "clash.pdf"}])
    (s.pending_dir / "clash.pdf").write_bytes(b"OLD")
    (s.input_dir / "clash.pdf").write_bytes(b"NEW")
    report = scanner.scan_and_stage(dry_run=False)
    assert report.renamed == 1
    assert report.staged == 1
    manifest = manifest_store.load(user_path)
    row = manifest["clash.pdf"]
    assert row.import_status == "已移入待处理(重命名)"
    assert row.process_status == "已扫描(重命名)"
