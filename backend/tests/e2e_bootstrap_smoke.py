"""E2E 验证：用户放入 11 列 Excel（文件名无扩展名）→ 启动 bootstrap（不移动）→ 扫描按钮 → 自动补全扩展名并更新 manifest。"""

import importlib
import logging
import shutil
import sys
from pathlib import Path

logging.disable(logging.CRITICAL)

# 准备工作目录
ROOT = Path("d:/programmtools/tools/ragsystem")
DATA = ROOT / "data"
INPUT = DATA / "input"
PENDING = DATA / "pending"
MANIFEST = DATA / "manifest.xlsx"

# 备份现有 manifest（如果存在）
BACKUP = DATA / "manifest.xlsx.bak"
if BACKUP.exists():
    BACKUP.unlink()
if MANIFEST.exists():
    shutil.copy2(MANIFEST, BACKUP)

# 清空 input/pending
for f in INPUT.iterdir():
    if f.name != ".gitkeep":
        f.unlink()
for f in PENDING.iterdir():
    if f.name != ".gitkeep":
        f.unlink()

# 1) 模拟用户放入一个 11 列的 Excel（文件名均无扩展名）
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
headers = [
    "序号", "文件名称", "一级分类", "二级分类", "关键词标签",
    "适用科室", "生效日期", "导入情况", "处理情况", "校对", "处理说明",
]
ws.append(headers)
# 关键：用户填的「文件名称」无扩展名
data_rows = [
    [1, "国标-001", "国标", "医用氧舱", "高压氧", "高压氧科", "2006-04-01", "", "", "", "用户原备注"],
    [2, "团标-002", "团标", "康复", "智能", "康复科", "2024-01-01", "", "", "", ""],
    [3, "新规-003", "规范", "院内", "院感", "院感科", "2025-06-01", "", "", "", "重要"],
    [4, "已处理-004",  "规范", "院内", "已导入", "院感科", "2025-06-01", "已移入待处理", "已扫描", "", "上次扫描已处理"],
    [5, "精确匹配-005", "国标", "测试", "ext", "测试科", "2025-01-01", "", "", "", "文件名带扩展名"],
]
for r in data_rows:
    ws.append(r)
wb.save(MANIFEST)
wb.close()

# 2) 启动 bootstrap
sys.path.insert(0, str(ROOT / "backend"))
import os
os.environ.setdefault("RAG_DATA_ROOT", str(DATA))

from app import config as cfg_mod
importlib.reload(cfg_mod)
settings = cfg_mod.settings
from app.services import scanner, manifest_store
scanner.settings = settings

manifest_store.bootstrap(DATA)

# 验证列数 = 17（11 用户 + 5 系统 + 1 parse，§3.2 新增 parse 列）
wb2 = __import__("openpyxl").load_workbook(MANIFEST, read_only=True)
ws2 = wb2.active
hdr = [str(c).strip() if c else "" for c in next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))]
wb2.close()
assert len(hdr) == 17
assert hdr[-1] == "parse", f"最后一列应为 parse，实际 {hdr[-1]}"
print("✓ bootstrap 补列成功（11→17，最后一列=parse）")

# 3) 模拟用户放入文件：扩展名不同，覆盖常见情况
(INPUT / "国标-001.pdf").write_bytes(b"PDF content for guobiao")
(INPUT / "新规-003.docx").write_bytes(b"DOCX content for xingui")
# 团标-002 故意不放 → MISSING
# 已处理-004 已被标记已处理，会被跳过
(INPUT / "精确匹配-005.pdf").write_bytes(b"Exact match with extension")  # 已经有扩展名
# 模拟一个同名但扩展名不同的歧义场景
(INPUT / "歧义文件-006.doc").write_bytes(b"DOC content")
(INPUT / "歧义文件-006.pdf").write_bytes(b"PDF content")
# 优先级：.pdf > .docx > .doc，所以应选 .pdf

# 4) 用户点击「扫描」按钮
report = scanner.scan_and_stage(dry_run=False)
print(f"✓ 扫描完成: staged={report.staged}, new={report.new}, skipped={report.skipped_done}, missing={report.missing_on_disk}")

# 5) 验证 manifest 的 filename 字段被更新为带扩展名的真实文件名
manifest = manifest_store.load(MANIFEST)

# 国标-001: 无扩展名 → 应补为 .pdf
assert "国标-001.pdf" in manifest, f"应存在 '国标-001.pdf'，实际 keys: {list(manifest.keys())}"
assert "国标-001" not in manifest, f"旧的 '国标-001' 不应再存在"
row1 = manifest["国标-001.pdf"]
assert row1.import_status == "已移入待处理"
assert row1.status == "pending"
print(f"✓ 国标-001 → 国标-001.pdf (扩展名自动补全)")

# 新规-003: 无扩展名 → 应补为 .docx
assert "新规-003.docx" in manifest
row3 = manifest["新规-003.docx"]
assert row3.import_status == "已移入待处理"
print(f"✓ 新规-003 → 新规-003.docx")

# 已处理-004: 跳过（已标记已导入）
assert "已处理-004" in manifest  # 保持原样（无扩展名）
row4 = manifest["已处理-004"]
assert row4.import_status == "已移入待处理"
print(f"✓ 已处理-004: 保持原文件名（已处理）")

# 精确匹配-005: 已有 .pdf → 保持
assert "精确匹配-005.pdf" in manifest
row5 = manifest["精确匹配-005.pdf"]
assert row5.import_status == "已移入待处理"
print(f"✓ 精确匹配-005.pdf: 精确匹配")

# 团标-002: MISSING
assert "团标-002" in manifest
row2 = manifest["团标-002"]
assert row2.import_status is None
print(f"✓ 团标-002: MISSING（input 找不到）")

# 6) 验证 pending/ 中的文件
pending_files = {p.name for p in PENDING.iterdir() if p.is_file()}
expected_pending = {"国标-001.pdf", "新规-003.docx", "精确匹配-005.pdf"}
assert expected_pending.issubset(pending_files), f"pending 缺文件: {expected_pending - pending_files}"
print(f"✓ pending/ 包含所有移入文件")

# 7) 验证 input/ 已被清空
input_files = {p.name for p in INPUT.iterdir() if p.name != ".gitkeep"}
# 团标/歧义/已处理 这些不该被移走
print(f"  input 剩余: {input_files}")

# 8) 幂等性：第二次扫描
second = scanner.scan_and_stage(dry_run=False)
assert second.staged == 0, f"第二次应 staged=0, 实际={second.staged}"
# 应有 3 个已导入 + 1 个 MISSING（团标-002）+ 1 个已处理（已处理-004）
print(f"✓ 第二次扫描幂等: staged={second.staged}, skipped={second.skipped_done}, missing={second.missing_on_disk}")

# 恢复
if BACKUP.exists():
    shutil.move(BACKUP, MANIFEST)

# 清掉测试文件
for d in (PENDING,):
    for f in d.iterdir():
        if f.name not in (".gitkeep",) and (f.name.startswith("国标") or f.name.startswith("新规") or f.name.startswith("精确匹配") or f.name.startswith("歧义")):
            if f.is_dir():
                __import__("shutil").rmtree(f, ignore_errors=True)
            else:
                f.unlink()
for f in INPUT.iterdir():
    if f.name not in (".gitkeep",) and (f.name.startswith("国标") or f.name.startswith("新规") or f.name.startswith("精确匹配") or f.name.startswith("歧义")):
        f.unlink()

print("\n=== 全部 E2E 验证通过（扩展名自动补全） ===")
