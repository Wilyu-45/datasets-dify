"""
清理 manifest：移除重复条目和不在 pending/ 中的条目。
"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"d:\programmtools\tools\ragsystem\data")
PENDING = DATA_DIR / "pending"
MANIFEST = DATA_DIR / "manifest.xlsx"

# 1) 读取 manifest
wb = openpyxl.load_workbook(str(MANIFEST))
ws = wb.active
headers = [c.value for c in ws[1]]

fname_col = headers.index("文件名称") if "文件名称" in headers else None
status_col = headers.index("status") if "status" in headers else None
dify_col = headers.index("dify_doc_id") if "dify_doc_id" in headers else None

# 2) 收集所有条目
entries = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    fn = row[fname_col] if fname_col is not None else None
    if not fn or not str(fn).strip():
        continue
    status = str(row[status_col]).strip() if status_col is not None and row[status_col] else ""
    dify_id = str(row[dify_col]).strip() if dify_col is not None and row[dify_col] else ""
    entries.append({
        "row_idx": row_idx,
        "filename": str(fn).strip(),
        "status": status,
        "dify_id": dify_id,
        "row_data": list(row),
    })

print(f"原始条目数: {len(entries)}")

# 3) 获取 pending/ 中的文件
pending_files = set()
for f in PENDING.iterdir():
    if f.is_file() and f.name != ".gitkeep":
        pending_files.add(f.name)

print(f"pending/ 文件数: {len(pending_files)}")

# 4) 过滤：保留在 pending/ 中的条目，去重
seen_filenames = set()
filtered_entries = []

for e in entries:
    fn = e["filename"]
    
    # 跳过不在 pending/ 中的条目
    if fn not in pending_files:
        print(f"  移除 (不在 pending/): {fn}")
        continue
    
    # 跳过重复条目（保留第一个）
    if fn in seen_filenames:
        print(f"  移除 (重复): {fn}")
        continue
    
    seen_filenames.add(fn)
    filtered_entries.append(e)

print(f"\n过滤后条目数: {len(filtered_entries)}")

# 5) 重写 manifest
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.append(headers)

for e in filtered_entries:
    ws_new.append(e["row_data"])

wb_new.save(str(MANIFEST))
wb_new.close()

print(f"\n已清理 manifest.xlsx")
print(f"最终条目数: {len(filtered_entries)}")

# 6) 显示统计
from collections import Counter
status_counts = Counter(e["status"] for e in filtered_entries)
print(f"\n状态分布:")
for status, count in status_counts.items():
    print(f"  {status or '(empty)'}: {count}")

dify_counts = sum(1 for e in filtered_entries if e["dify_id"] and e["dify_id"] != "None")
print(f"\n有 Dify ID: {dify_counts}")
print(f"无 Dify ID: {len(filtered_entries) - dify_counts}")
