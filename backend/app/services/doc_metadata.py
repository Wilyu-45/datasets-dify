"""文档元数据管理（excel.txt → Dify Metadata API）。

从 Excel 文件读取文档元数据（按文件名 stem 关联），在 Dify 入库时：
    1. 确保知识库中存在对应的元数据字段（首次运行时自动创建）
    2. 文档入库成功后，批量设置该文档的元数据值

Excel 格式约定（excel.txt）：
    - 第 1 行：英文字段名（doc_type_primary, doc_type_secondary, ...）
    - 第 2 行：中文列名（类型-一级, 类型-二级, ...）— 仅作展示，代码用英文列名
    - 第 3 行起：数据行，第 1 列为文件 stem（不含后缀）

元数据字段定义：
    doc_type_primary    → string  (类型-一级)
    doc_type_secondary  → string  (类型-二级)
    topic_primary       → string  (主题-一级)
    topic_secondary     → string  (主题-二级)
    core_summary        → string  (核心内容摘要)
    entity_label        → string  (实体标签)
    attribute_label     → string  (属性标签)
    applicable_scenarios→ string  (适用科室)
    effective_date      → string  (生效日期 — 用 string 兼容 "无" / "2009-06-01" 混合格式)
    priority            → number  (优先级)
    status              → string  (现行/废止/...)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from app.config import settings

log = logging.getLogger("ragsystem.doc_metadata")


# ============ 字段定义 ============

# 英文字段名 → Dify 元数据 type
# 注意：effective_date 用 string 而非 time，因为用户 Excel 里有 "无" 这种非日期值
METADATA_FIELD_DEFS: Dict[str, str] = {
    "doc_type_primary": "string",
    "doc_type_secondary": "string",
    "topic_primary": "string",
    "topic_secondary": "string",
    "core_summary": "string",
    "entity_label": "string",
    "attribute_label": "string",
    "applicable_scenarios": "string",
    "effective_date": "string",
    "priority": "number",
    "status": "string",
}


# ============ Excel 读取 ============


def load_doc_metadata(excel_path: Optional[Path] = None) -> Dict[str, Dict[str, Any]]:
    """读取文档元数据 Excel。

    Returns:
        {stem: {field_name: value, ...}, ...}
        stem = 第一列值（strip 后），field_name = 英文列名
    """
    path = Path(excel_path or settings.doc_metadata_excel_path)
    if not path.exists():
        log.warning("文档元数据 Excel 不存在: %s", path)
        return {}

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # 第 1 行：英文字段名
        header_row = next(rows_iter, None)
        if header_row is None:
            return {}
        col_names = [str(c).strip() if c else "" for c in header_row]
        # 第一列是 stem（可能为空），至少需要有其他列
        if len(col_names) < 2:
            return {}

        # 第 2 行是中文列名，跳过
        try:
            next(rows_iter)
        except StopIteration:
            return {}

        # 数据行
        result: Dict[str, Dict[str, Any]] = {}
        for raw in rows_iter:
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            # 第 1 列 = stem
            stem_val = raw[0]
            if stem_val is None:
                continue
            stem = str(stem_val).strip()
            if not stem:
                continue

            row_data: Dict[str, Any] = {}
            for i, col_name in enumerate(col_names):
                if i >= len(raw):
                    break
                if i == 0:
                    continue  # 跳过 stem 列
                if not col_name or col_name not in METADATA_FIELD_DEFS:
                    continue
                val = raw[i]
                if val is None or str(val).strip() == "":
                    continue
                # 按字段类型转换
                field_type = METADATA_FIELD_DEFS[col_name]
                if field_type == "number":
                    try:
                        row_data[col_name] = float(val) if not isinstance(val, (int, float)) else val
                    except (TypeError, ValueError):
                        continue
                else:
                    row_data[col_name] = str(val).strip()

            if row_data:
                result[stem] = row_data

        log.info(
            "文档元数据 Excel 加载完成: path=%s docs=%d",
            path, len(result),
        )
        return result
    finally:
        wb.close()


# ============ Dify 元数据字段同步 ============


def ensure_metadata_fields(client: Any) -> Dict[str, Dict[str, Any]]:
    """确保 Dify 知识库中存在所有元数据字段。

    已存在的字段跳过（不重复创建），缺失的自动创建。

    Returns:
        {field_name: {id, name, type}, ...}  — 所有字段的 Dify 映射
    """
    from app.services.dify_uploader import DifyError

    # 1) 获取已有字段
    existing = client.list_metadata_fields()
    existing_by_name: Dict[str, Dict[str, Any]] = {
        f.get("name", ""): f for f in existing if f.get("name")
    }

    result: Dict[str, Dict[str, Any]] = {}

    # 2) 逐个检查 / 创建
    for field_name, field_type in METADATA_FIELD_DEFS.items():
        if field_name in existing_by_name:
            result[field_name] = existing_by_name[field_name]
            log.debug("元数据字段已存在: %s (id=%s)", field_name, result[field_name].get("id"))
            continue
        try:
            created = client.create_metadata_field(field_name, field_type)
            result[field_name] = created
            log.info(
                "元数据字段创建成功: %s (id=%s, type=%s)",
                field_name, created.get("id"), created.get("type"),
            )
        except DifyError as e:
            # 400 "already exists" → 并发创建场景，重新获取
            if e.status_code == 400 and "already exists" in (e.body or "").lower():
                fields = client.list_metadata_fields()
                for f in fields:
                    if f.get("name") == field_name:
                        result[field_name] = f
                        break
                if field_name not in result:
                    log.error("元数据字段创建报已存在但列表找不到: %s", field_name)
            else:
                log.error("元数据字段创建失败: %s — %s", field_name, e)

    return result


# ============ 文档元数据设置 ============


def build_metadata_operation(
    document_id: str,
    stem: str,
    field_map: Dict[str, Dict[str, Any]],
    doc_meta: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """为单个文档构建 metadata batch update 的 operation_data 项。

    Args:
        document_id: Dify 文档 ID
        stem: 文档 stem（用于在 doc_meta 中查找）
        field_map: ensure_metadata_fields 返回的 {field_name: {id, name, type}}
        doc_meta: load_doc_metadata 返回的 {stem: {field: value}}

    Returns:
        operation_data 的一项，或 None（无匹配元数据时）
    """
    meta_values = doc_meta.get(stem)
    if not meta_values:
        return None

    metadata_list: List[Dict[str, Any]] = []
    for field_name, value in meta_values.items():
        field_info = field_map.get(field_name)
        if not field_info:
            continue
        metadata_list.append({
            "id": field_info.get("id", ""),
            "name": field_name,
            "value": value,
        })

    if not metadata_list:
        return None

    return {
        "document_id": document_id,
        "metadata_list": metadata_list,
        "partial_update": True,
    }
