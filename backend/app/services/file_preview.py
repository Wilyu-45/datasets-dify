"""落地文件在线预览（2026-09 新增，网页抓取「下载后文件预览」使用）。

设计目标：能看的格式直接看、不能看的给文件信息。
    - pdf            → 浏览器原生 PDF 查看器（iframe /file）
    - html/htm       → 后端清理脚本等危险内容后 iframe（/file）
    - png/jpg/...    → <img>（/file）
    - md/txt         → 后端解码为 UTF-8 文本（/file 返回 text/plain，前端渲染）
    - docx/pptx      → 基于 OOXML(zip+xml) 的轻量文本/表格/逐页提取（/office-preview 返回 HTML）
    - xlsx/csv       → 基于 openpyxl / csv 渲染为表格 HTML（/office-preview）
    - .doc/.xls/.ppt 旧版 Office → 经本机 LibreOffice 无头转换为新格式后按上两类预览
      （自动探测 soffice；未安装/转换失败时退回信息页，可下载自查）
    - 压缩包/邮件等无法转换的二进制 → 文件信息页（可下载自查）
"""

from __future__ import annotations

import contextlib
import csv
import html as html_mod
import io
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import zipfile
from collections import OrderedDict
from pathlib import Path
from xml.etree import ElementTree as ET

log = logging.getLogger("ragsystem.file_preview")

# ---- 类型判定 ----

_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
_HTML_EXTS = {".html", ".htm"}
_TEXT_EXTS = {".txt", ".md", ".markdown"}
_ARCHIVE_EXTS = {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz"}
# 旧版二进制 Office：无纯 Python 解析方案，先经 LibreOffice 无头转换为新格式
_LEGACY_TARGET_EXT = {".doc": "docx", ".xls": "xlsx", ".ppt": "pptx"}
_LEGACY_TARGET_LABEL = {
    ".doc": "Word 97-2003 (.doc)",
    ".xls": "Excel 97-2003 (.xls)",
    ".ppt": "PowerPoint 97-2003 (.ppt)",
}


def preview_kind(filename: str) -> str:
    """按扩展名给出预览类型。

    取值：pdf / image / html / markdown / text / docx / xlsx / pptx / csv /
          doc / xls / ppt / archive / other
    """
    ext = Path(filename or "").suffix.lower()
    if ext == ".pdf":
        return "pdf"
    if ext in _IMAGE_EXTS:
        return "image"
    if ext in _HTML_EXTS:
        return "html"
    if ext == ".md" or ext == ".markdown":
        return "markdown"
    if ext in _TEXT_EXTS:
        return "text"
    if ext in (".docx",):
        return "docx"
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    if ext == ".pptx":
        return "pptx"
    if ext == ".csv":
        return "csv"
    if ext in (".doc",):
        return "doc"
    if ext == ".xls":
        return "xls"
    if ext == ".ppt":
        return "ppt"
    if ext in _ARCHIVE_EXTS:
        return "archive"
    return "other"


def decode_bytes(data: bytes) -> str:
    """字节解码为文本：优先 UTF-8，失败回退 GB18030（中文政府站点常见）。"""
    for enc in ("utf-8", "gb18030"):
        try:
            return data.decode(enc)
        except (UnicodeDecodeError, ValueError):
            continue
    return data.decode("utf-8", errors="replace")


# ---- HTML 清洗（第三方 HTML 原样预览时使用，防脚本/跳转） ----


def sanitize_html(text: str) -> str:
    """移除可执行/干扰元素与属性：script/style/iframe/object/embed/form 等、
    on* 事件、javascript: 伪协议、meta refresh、base。"""
    # 成对标签整体删除（含内容）
    text = re.sub(
        r"<\s*(script|style|iframe|frame|object|embed|applet|form|frameset|base)[^>]*>.*?<\s*/\s*\1\s*>",
        "",
        text,
        flags=re.I | re.S,
    )
    # 残留的开始/自闭合危险标签
    text = re.sub(
        r"<\s*(script|style|iframe|frame|object|embed|applet|form|frameset|base|meta)[^>]*/?>",
        "",
        text,
        flags=re.I,
    )
    # 事件属性 onxxx="..."
    text = re.sub(
        r'\s+on\w+\s*=\s*(?:"[^"]*"|\'[^\']*\'|[^\s>]+)',
        "",
        text,
        flags=re.I,
    )
    # javascript: / vbscript: 伪协议
    text = re.sub(
        r'\s+(?:href|src|action|data|formaction)\s*=\s*["\']?\s*(?:javascript|vbscript):[^"\'>\s]*',
        "",
        text,
        flags=re.I,
    )
    return text


def cleaned_html_bytes(data: bytes) -> bytes:
    """HTML 文件预览：解码 → 清洗 → UTF-8 字节。"""
    return sanitize_html(decode_bytes(data)).encode("utf-8", errors="replace")


# ---- Office/表格 → HTML 轻量预览（后端生成，前端 iframe） ----


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _attr_val(el: ET.Element, local_name: str) -> str:
    """取属性（兼容带命名空间）：如 gridSpan 的 w:val。"""
    for k, v in el.attrib.items():
        if _local(k) == local_name:
            return v or ""
    return ""


def _info_page_html(title: str, reason: str) -> str:
    body = (
        '<div class="pv-info">'
        f"<div class='pv-info-icon'>&#128196;</div>"
        f"<div class='pv-info-title'>{html_mod.escape(title)}</div>"
        f"<p>{html_mod.escape(reason)}</p>"
        "<p>该文件为本次确认下载的原文件，已保存到待处理区；"
        "可直接使用下方「下载原文件」在本地打开查看后再回来点「确定」入库。</p>"
        "</div>"
    )
    return _wrap_html(title, body)


# ---------- DOCX（zip + 主文档 XML，段落/表格/文本，图片以占位符标记） ----------


def _docx_para_html(p: ET.Element) -> str:
    parts: list[str] = []
    has_img = False
    for node in p.iter():
        t = _local(node.tag)
        if t == "t":
            parts.append(html_mod.escape(node.text or ""))
        elif t in ("br", "cr"):
            parts.append("<br>")
        elif t == "tab":
            parts.append("&nbsp;&nbsp;")
        elif t in ("drawing", "pict", "object", "embeddedObject"):
            has_img = True
    text = "".join(parts)
    if not text.strip() and not has_img:
        return ""
    if has_img:
        text = f"{text} <span class='ph'>[图片]</span>".strip()
    return f"<p>{text}</p>"


def _docx_table_html(tbl: ET.Element) -> str:
    rows: list[str] = []
    for tr in tbl:
        if _local(tr.tag) != "tr":
            continue
        cells: list[str] = []
        for tc in tr:
            if _local(tc.tag) != "tc":
                continue
            colspan = 1
            for sub in tc:
                if _local(sub.tag) != "tcPr":
                    continue
                for prop in sub:
                    if _local(prop.tag) == "gridSpan":
                        try:
                            colspan = max(1, int(_attr_val(prop, "val") or "1"))
                        except ValueError:
                            colspan = 1
            inner: list[str] = []
            for cc in tc:
                t = _local(cc.tag)
                if t == "p":
                    inner.append(_docx_para_html(cc))
                elif t == "tbl":
                    inner.append(_docx_table_html(cc))
            cells.append(f"<td colspan='{colspan}'>{''.join(inner)}</td>")
        rows.append(f"<tr>{''.join(cells)}</tr>")
    if not rows:
        return ""
    return f"<table class='pv-table'><tbody>{''.join(rows)}</tbody></table>"


def _docx_block_parts(el: ET.Element, out: list[str]) -> None:
    for child in el:
        t = _local(child.tag)
        if t == "p":
            out.append(_docx_para_html(child))
        elif t == "tbl":
            out.append(_docx_table_html(child))
        elif t == "sdt":  # 结构文档标签，内容在 sdtContent
            for inner in child:
                if _local(inner.tag) == "sdtContent":
                    _docx_block_parts(inner, out)


def docx_to_html(data: bytes) -> str:
    """提取 docx 主文档文本为 HTML（段落 + 表格；图片给占位符）。"""
    try:
        with io.BytesIO(data) as buf, zipfile.ZipFile(buf) as zf:
            xml = zf.read("word/document.xml")
    except Exception:  # noqa: BLE001 非标准 docx
        return _info_page_html("Word 文档", "该 .docx 结构无法解析，暂不能在线预览。")
    try:
        root = ET.fromstring(xml)
    except ET.ParseError:
        return _info_page_html("Word 文档", "该 .docx 结构无法解析，暂不能在线预览。")
    parts: list[str] = []
    for child in root:
        if _local(child.tag) == "body":
            _docx_block_parts(child, parts)
    body = "".join(x for x in parts if x)
    if not body:
        body = "<p>（未在文档中提取到文字内容，可能为纯图片/扫描件）</p>"
    return _wrap_html("Word 文档预览", body)


# ---------- PPTX（zip + 每页 slide XML 的文本段落） ----------


def _pptx_collect_text(node: ET.Element, out: list[str]) -> None:
    t = _local(node.tag)
    if t == "t":
        out.append(node.text or "")
        return
    if t == "br":
        out.append("\n")
        return
    if t == "tab":
        out.append("\t")
        return
    if t == "p":
        out.append("\n")
    for c in node:
        _pptx_collect_text(c, out)


def pptx_to_html(data: bytes) -> str:
    """把 pptx 每页文字按幻灯片输出为 HTML（图片/版式不强求还原）。"""
    try:
        with io.BytesIO(data) as buf, zipfile.ZipFile(buf) as zf:
            slide_names = sorted(
                (
                    n
                    for n in zf.namelist()
                    if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)
                ),
                key=lambda n: int(re.search(r"(\d+)", n).group(1)),
            )
            slides: list[str] = []
            for name in slide_names:
                num = int(re.search(r"(\d+)", name).group(1))
                try:
                    root = ET.fromstring(zf.read(name))
                except ET.ParseError:
                    continue
                out: list[str] = []
                _pptx_collect_text(root, out)
                text = "".join(out)
                lines = [ln.strip() for ln in text.split("\n")]
                text = "\n".join(ln for ln in lines if ln).strip()
                if text:
                    body = html_mod.escape(text).replace("\n", "<br>")
                else:
                    body = "<p class='muted'>（本页无文字内容，可能为图片或图表）</p>"
                slides.append(
                    f"<div class='slide'><div class='slide-no'>第 {num} 张幻灯片</div>"
                    f"<div class='slide-body'>{body}</div></div>"
                )
    except Exception:  # noqa: BLE001 非标准 pptx
        return _info_page_html("PPT 演示文稿", "该 .pptx 结构无法解析，暂不能在线预览。")
    if not slides:
        return _info_page_html("PPT 演示文稿", "未在演示文稿中提取到可预览内容。")
    return _wrap_html("PPT 演示文稿预览", "".join(slides))


# ---------- XLSX / CSV（表格 HTML） ----------


def _cell_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 300:
        s = s[:300] + "…"
    return s


def xlsx_to_html(path: Path) -> str:
    """openpyxl 读取所有工作表 → 表格 HTML（只读，限制行列避免超大文件拖垮预览）。"""
    try:
        from openpyxl import load_workbook
    except Exception as e:  # noqa: BLE001
        return _info_page_html("Excel 工作簿", f"缺少 openpyxl，暂不能在线预览（{e}）。")
    sections: list[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for idx, ws in enumerate(wb.worksheets[:50]):
            rows_html: list[str] = []
            total_rows = 0
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                if total_rows > 2000:
                    break
                cells = "".join(
                    f"<td>{html_mod.escape(_cell_text(v))}</td>"
                    for v in row[:80]
                )
                rows_html.append(f"<tr>{cells}</tr>")
            head = f"<div class='sheet'><div class='sheet-name'>{html_mod.escape(ws.title)}</div>"
            if not rows_html:
                head += "<p class='muted'>（工作表为空）</p></div>"
            else:
                head += (
                    "<div class='table-scroll'><table class='pv-table'>"
                    f"<tbody>{''.join(rows_html)}</tbody></table></div></div>"
                )
            if total_rows > 2000:
                head = head.replace(
                    "</div></div>",
                    "<p class='muted'>（仅预览前 2000 行，完整内容见原文件）</p></div></div>",
                    1,
                )
            sections.append(head)
        wb.close()
    except Exception:  # noqa: BLE001
        return _info_page_html("Excel 工作簿", "该 Excel 文件无法解析，暂不能在线预览。")
    if not sections:
        return _info_page_html("Excel 工作簿", "工作簿中没有可预览的工作表。")
    return _wrap_html("Excel 工作簿预览", "".join(sections))


def csv_to_html(data: bytes) -> str:
    text = decode_bytes(data)
    try:
        reader = list(csv.reader(io.StringIO(text)))
    except Exception:  # noqa: BLE001
        reader = []
    if not reader:
        return _wrap_html("CSV 文件预览", "<p>（CSV 无可预览内容）</p>")
    total = len(reader)
    rows_html: list[str] = []
    for row in reader[:1000]:
        cells = "".join(
            f"<td>{html_mod.escape(_cell_text(v))}</td>" for v in row[:80]
        )
        rows_html.append(f"<tr>{cells}</tr>")
    note = f"<p class='muted'>共 {total} 行，展示前 {min(1000, total)} 行。</p>" if total > 1000 else ""
    return _wrap_html(
        "CSV 文件预览",
        note + f"<div class='table-scroll'><table class='pv-table'><tbody>{''.join(rows_html)}</tbody></table></div>",
    )


# ---------- 旧版 Office（.doc/.xls/.ppt）：LibreOffice 无头转换 ----------


def _configured_soffice() -> str:
    """配置指定/禁用：RAG_OFFICE_SOFFICE_PATH（支持 backend/.env 与环境变量）。"""
    try:
        from app.config import settings  # 惰性：保持本模块可独立自测

        raw = (settings.office_soffice_path or "").strip()
    except Exception:  # noqa: BLE001 独立运行/自测环境无项目配置时退回环境变量
        raw = ""
    if not raw:
        raw = os.environ.get("RAG_OFFICE_SOFFICE_PATH", "").strip()
    return raw


_detected_soffice: str | None = None  # None=未探测；''=确认未安装；否则为可执行文件路径


def _normalize_soffice(p: Path) -> str:
    """Windows 下 soffice.exe/.com 是启动壳，部分环境会挂起/误判；
    同目录存在真正的引擎 soffice.bin 时优先用它（无头转换更稳）。"""
    if not p.is_file():
        return ""
    if os.name == "nt" and p.name.lower() in ("soffice.exe", "soffice.com"):
        sibling = p.parent / "soffice.bin"
        if sibling.is_file():
            return str(sibling)
    return str(p)


def _detect_soffice() -> str:
    """探测可用的 soffice：配置 → PATH → 常见安装目录。结果进程内缓存。"""
    global _detected_soffice
    if _detected_soffice is not None:
        return _detected_soffice
    result = ""
    cfg = _configured_soffice()
    if cfg.lower() not in ("none", "off", "0"):
        if cfg:
            result = _normalize_soffice(Path(cfg))
        if not result:
            for name in ("soffice", "soffice.exe", "soffice.bin"):
                hit = shutil.which(name)
                if hit:
                    result = _normalize_soffice(Path(hit))
                    break
        if not result:
            cands: list[Path] = []
            if os.name == "nt":
                local = os.environ.get("LOCALAPPDATA")
                cands = [
                    Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
                    Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
                ]
                if local:
                    cands.append(Path(local) / "Programs" / "LibreOffice" / "program" / "soffice.exe")
            elif sys.platform == "darwin":
                cands = [Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")]
            else:
                cands = [Path("/usr/bin/soffice"), Path("/usr/lib/libreoffice/program/soffice")]
            for c in cands:
                result = _normalize_soffice(c)
                if result:
                    break
    if result:
        log.info("网页抓取旧版 Office 预览使用 LibreOffice: %s", result)
    _detected_soffice = result
    return result


_SOFFICE_LOCK = threading.Lock()  # LibreOffice 单实例安全：串行执行转换


@contextlib.contextmanager
def _soffice_outdir(src: Path, target_ext: str):
    """执行 LibreOffice 无头转换（doc/xls/ppt → docx/xlsx/pptx）。

    yield 临时目录（转换产物位于 {tmp}/{src.stem}.{target_ext}），退出时清理；
    转换不可用时 yield None。
    注意：Windows 上 soffice.bin 使用 -env:UserInstallation 临时 profile 反而会
    启动失败（rc=81），因此不设置，改用全局锁 _SOFFICE_LOCK 串行执行，避免与
    其它并发转换争用同一用户 profile。
    """
    soffice = _detect_soffice()
    if not soffice:
        yield None
        return
    work = Path(tempfile.mkdtemp(prefix="rag_office_pv_"))
    try:
        cmd = [
            soffice,
            "--headless",
            "--norestore",
            "--nolockcheck",
            "--convert-to",
            target_ext,
            "--outdir",
            str(work),
            str(src),
        ]
        # LibreOffice 内部带 Python 运行时：必须清除继承的 PYTHON*/VIRTUAL_ENV，
        # 否则（如 IDE 注入的 PYTHONPATH）会导致 soffice 启动即失败/卡死。
        clean_env = {
            k: v for k, v in os.environ.items()
            if not k.startswith("PYTHON") and k != "VIRTUAL_ENV"
        }
        try:
            with _SOFFICE_LOCK:
                proc = subprocess.run(cmd, capture_output=True, timeout=240, env=clean_env)
        except Exception as e:  # noqa: BLE001
            log.warning("LibreOffice 转换调用异常: file=%s err=%s", src.name, e)
            yield None
            return
        if proc.returncode != 0:
            tail = (proc.stderr or b"").decode("utf-8", "replace").strip()[-300:]
            log.warning("LibreOffice 转换失败: file=%s rc=%s err=%s", src.name, proc.returncode, tail)
            yield None
            return
        yield work
    finally:
        shutil.rmtree(work, ignore_errors=True)


def _legacy_office_html(filename: str, path: Path) -> str:
    """旧版 Office → LibreOffice 转 docx/xlsx/pptx → 复用新格式 HTML 预览。"""
    ext = Path(filename or "").suffix.lower()
    target_ext = _LEGACY_TARGET_EXT.get(ext)
    label = _LEGACY_TARGET_LABEL.get(ext, "旧版 Office 文档")
    if not target_ext:
        return _info_page_html(label, "暂不支持该格式在线预览，请下载原文件用本地软件查看。")
    if not _detect_soffice():
        return _info_page_html(
            label,
            "本机未检测到 LibreOffice，旧版 Office 暂无法在线预览。"
            "可先「下载原文件」用本地 Office/WPS 查看；"
            "如需在线预览，请在服务器安装 LibreOffice 后重启服务。",
        )
    with _soffice_outdir(path, target_ext) as work:
        if work is None:
            return _info_page_html(
                label,
                "LibreOffice 转换失败，旧版 Office 暂无法在线预览；"
                "可先「下载原文件」用本地 Office/WPS 查看。",
            )
        out = work / f"{path.stem}.{target_ext}"
        if not out.is_file():
            return _info_page_html(label, "LibreOffice 未产出转换结果，暂无法在线预览。")
        try:
            if target_ext == "xlsx":
                return xlsx_to_html(out)
            data = out.read_bytes()
            return docx_to_html(data) if target_ext == "docx" else pptx_to_html(data)
        except Exception as e:  # noqa: BLE001
            log.exception("旧版 Office 转换后预览失败: file=%s", path.name)
            return _info_page_html(label, f"转换后预览生成失败：{e}；可先下载原文件查看。")


# ---------- 预览结果缓存（旧版转换耗时，同一文件重复预览直接命中） ----------

_PREVIEW_CACHE: "OrderedDict[tuple, str]" = OrderedDict()
_PREVIEW_CACHE_LOCK = threading.Lock()
_PREVIEW_CACHE_MAX = 24


def _preview_cache_key(filename: str, path: Path):
    try:
        st = path.stat()
    except OSError:
        return None
    return (filename, str(path.resolve()), st.st_size, st.st_mtime_ns)


def _preview_cache_get(filename: str, path: Path) -> str | None:
    key = _preview_cache_key(filename, path)
    if key is None:
        return None
    with _PREVIEW_CACHE_LOCK:
        html = _PREVIEW_CACHE.get(key)
        if html is not None:
            _PREVIEW_CACHE.move_to_end(key)
        return html


def _preview_cache_put(filename: str, path: Path, html: str) -> None:
    key = _preview_cache_key(filename, path)
    if key is None:
        return
    with _PREVIEW_CACHE_LOCK:
        _PREVIEW_CACHE[key] = html
        _PREVIEW_CACHE.move_to_end(key)
        while len(_PREVIEW_CACHE) > _PREVIEW_CACHE_MAX:
            _PREVIEW_CACHE.popitem(last=False)


# ---------- 汇总入口 + 通用 HTML 外壳 ----------


_PV_CSS = """
*{box-sizing:border-box}
body{margin:0;background:#fff;color:#24292f;
  font:14px/1.7 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif}
.pv{padding:20px 24px;word-break:break-word}
.pv h1{font-size:18px;margin:0 0 4px}
.pv .meta{color:#6b7280;font-size:12px;margin-bottom:12px}
.pv p{margin:6px 0}
.pv .muted{color:#6b7280}
.pv .ph{color:#b45309;font-size:12px;border:1px dashed #d9a460;border-radius:4px;padding:0 4px}
.pv .pv-info{text-align:center;color:#4b5563;padding:60px 30px}
.pv .pv-info-icon{font-size:48px;margin-bottom:10px}
.pv .pv-info-title{font-size:16px;font-weight:600;color:#111827;margin-bottom:6px}
.pv .pv-table{border-collapse:collapse;width:100%;margin:8px 0;font-size:13px}
.pv .pv-table td,.pv .pv-table th{border:1px solid #d0d7de;padding:4px 8px;vertical-align:top}
.table-scroll{overflow:auto;max-height:60vh}
.sheet{margin-bottom:14px}
.sheet-name{font-weight:600;color:#1f2328;margin:6px 0;border-left:3px solid #1677ff;padding-left:8px}
.slide{margin-bottom:16px;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden}
.slide-no{background:#f3f4f6;padding:4px 12px;font-weight:600;font-size:13px;color:#374151}
.slide-body{padding:10px 14px}
.slide-body p{margin:4px 0}
"""


def _wrap_html(title: str, body: str) -> str:
    return (
        "<!DOCTYPE html><html lang='zh'><head><meta charset='utf-8'>"
        f"<title>{html_mod.escape(title)}</title><style>{_PV_CSS}</style></head>"
        f"<body class='pv'>{body}</body></html>"
    )


def render_office_preview(filename: str, path: Path) -> str:
    """office-preview 接口入口：按扩展名生成可 iframe 的 HTML 预览页。

    .docx/.pptx 走 XML 提取，.xlsx/.csv 走表格渲染；
    .doc/.xls/.ppt 旧版 Office 先经 LibreOffice 无头转换为新格式再按上两类预览
    （找不到 soffice / 转换失败退回信息页）；其余二进制返回信息页。
    结果按文件内容状态缓存（旧版转换耗时数秒，避免同一文件重复转换）。
    """
    kind = preview_kind(filename)
    ext = Path(filename or "").suffix.lower()
    cached = _preview_cache_get(filename, path)
    if cached is not None:
        return cached
    if kind == "docx":
        try:
            html = docx_to_html(path.read_bytes())
        except Exception:  # noqa: BLE001
            html = _info_page_html("Word 文档", "读取文件失败，无法在线预览。")
    elif kind == "pptx":
        try:
            html = pptx_to_html(path.read_bytes())
        except Exception:  # noqa: BLE001
            html = _info_page_html("PPT 演示文稿", "读取文件失败，无法在线预览。")
    elif kind == "xlsx":
        html = xlsx_to_html(path)
    elif kind == "csv":
        try:
            html = csv_to_html(path.read_bytes())
        except Exception:  # noqa: BLE001
            html = _info_page_html("CSV 文件", "读取文件失败，无法在线预览。")
    elif kind in ("doc", "xls", "ppt"):
        html = _legacy_office_html(filename, path)
    else:
        # 压缩包等无法在线预览的二进制 → 文件信息页（抽屉另提供「下载原文件」按钮）
        friendly = _LEGACY_TARGET_LABEL.get(ext, f".{ext.lstrip('.')} 格式")
        html = _info_page_html(
            friendly,
            "该二进制格式暂不支持在浏览器中直接预览，"
            "点击下方「下载原文件」在本地打开查看。",
        )
    _preview_cache_put(filename, path, html)
    return html
