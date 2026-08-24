"""manifest.xlsx 的 openpyxl 包装。

列结构（与 plan.md 3.1/3.2/3.3 + 文件列表Excel示例 对齐）：
    用户原有 11 列（顺序可任意）：
       1. 序号
       2. 文件名称         ← 主键
       3. 一级分类
       4. 二级分类
       5. 关键词标签
       6. 适用科室
       7. 生效日期
       8. 导入情况
       9. 处理情况
      10. 校对
      11. 处理说明
    系统追加 5 列（始终在末尾）：
      12. status          (管线 FSM: new/pending/.../done/error)
      13. md5
      14. create_time
      15. update_time
      16. error_msg
    §3.2 追加 1 列（始终在末尾）：
      17. parse           (解析产物目录：data/parsed/{stem}/ 或错误描述)
    §3.3 追加 1 列（始终在末尾）：
      18. chunks          (切分产物目录：data/chunks/{stem}/ 或错误描述)
    §3.4 追加 2 列（始终在末尾）：
      19. dify_doc_id     (Dify 文档 ID)
      20. dify_status     (done / error / 空)

设计要点：
    - 用户可在 data/manifest.xlsx 中只放 11 列；启动时（ensure_columns）自动
      追加缺失的处理列到表尾，不动用户原有列及数据。
    - 读/写都按表头名匹配，不依赖位置；用户表头顺序任意。
    - "导入情况" 列（用户列）由系统更新以反映"已移入待处理"；
      "处理情况" 列由各步骤完成后更新以反映当前处理阶段。
    - "parse" 列（系统列）由 §3.2 步骤完成后更新。
    - "chunks" 列（系统列）由 §3.3 步骤完成后更新。
    - "dify_doc_id" / "dify_status"（系统列）由 §3.4 步骤完成后更新。
"""

from __future__ import annotations

import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.schemas import ManifestRow

log = logging.getLogger("ragsystem.manifest_store")


# ============ 列表头定义 ============

# 全部 20 列（用户原 11 + 系统 5 + §3.2 解析 1 + §3.3 切分 1 + §3.4 Dify 2）
HEADERS_ZH: List[str] = [
    "序号",
    "文件名称",
    "一级分类",
    "二级分类",
    "关键词标签",
    "适用科室",
    "生效日期",
    "导入情况",
    "处理情况",
    "校对",
    "处理说明",
    "status",
    "md5",
    "create_time",
    "update_time",
    "error_msg",
    "parse",
    "chunks",
    "dify_doc_id",
    "dify_status",
]

# 系统追加的 5 列（始终追加在用户表末尾）
SYSTEM_HEADERS_ZH: List[str] = ["status", "md5", "create_time", "update_time", "error_msg"]

# §3.2 追加的 1 列（始终在系统列之后）
PARSE_HEADER_ZH: str = "parse"

# §3.3 追加的 1 列（始终在 parse 列之后）
CHUNKS_HEADER_ZH: str = "chunks"

# §3.4 追加的 2 列（始终在 chunks 列之后）
DIFY_HEADERS_ZH: List[str] = ["dify_doc_id", "dify_status"]

# 用户原始 11 列（用于在写盘时定位"导入情况/处理情况"等列）
USER_HEADERS_ZH: List[str] = [
    h
    for h in HEADERS_ZH
    if h not in SYSTEM_HEADERS_ZH
    and h != PARSE_HEADER_ZH
    and h != CHUNKS_HEADER_ZH
    and h not in DIFY_HEADERS_ZH
]

# 表头列名 → ManifestRow 字段名
HEADER_TO_FIELD: Dict[str, str] = {
    "序号": "seq",
    "文件名称": "filename",
    "一级分类": "category_l1",
    "二级分类": "category_l2",
    "关键词标签": "keywords",
    "适用科室": "department",
    "生效日期": "effective_date",
    "导入情况": "import_status",
    "处理情况": "process_status",
    "校对": "verified",
    "处理说明": "process_note",
    "status": "status",
    "md5": "md5",
    "create_time": "create_time",
    "update_time": "update_time",
    "error_msg": "error_msg",
    "parse": "parse",
    "chunks": "chunks",
    "dify_doc_id": "dify_doc_id",
    "dify_status": "dify_status",
}

# ManifestRow 字段名 → 表头列名
FIELD_TO_HEADER: Dict[str, str] = {v: k for k, v in HEADER_TO_FIELD.items()}


# 模块级写锁：openpyxl 不是线程安全的，串行化所有写
_write_lock = threading.Lock()


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ============ 文件定位 ============


def find_manifest_file(data_dir: Path) -> Path:
    """返回 manifest 文件路径。固定为 data/manifest.xlsx。"""
    return data_dir / "manifest.xlsx"


# ============ 表头与初始化 ============


def ensure_exists(path: Path) -> None:
    """manifest 不存在时创建并写入 18 列表头。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    _write_headers(path, HEADERS_ZH)


def _write_headers(path: Path, headers: List[str]) -> None:
    """创建新工作簿并写入给定表头。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    _style_header(ws, len(headers))
    _set_column_widths(ws, headers)
    wb.save(path)


def _style_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(ws, headers: List[str]) -> None:
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(header) * 2)


def _backup_corrupted(path: Path) -> None:
    """把损坏的 manifest 文件备份为 {name}.corrupted.{timestamp}。

    避免直接删除，保留手动恢复的可能性。
    """
    import shutil

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_suffix(f".corrupted.{ts}.xlsx")
    try:
        shutil.move(str(path), str(backup))
        log.info("损坏文件已备份: %s", backup.name)
    except OSError as e:
        log.warning("无法备份损坏文件: %s", e)


def _read_header(path: Path) -> List[str]:
    """读首行表头（已 strip）。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        try:
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return []
        return [str(c).strip() if c is not None else "" for c in row]
    finally:
        wb.close()


def ensure_columns(path: Path) -> Tuple[bool, List[str]]:
    """确保所有 18 列存在。缺失的列追加在表尾。

    行为：
    - 文件不存在 → 创建 18 列；返回 (True, HEADERS_ZH)
    - 已含全部 18 列 → 不动；返回 (False, 当前表头)
    - 部分缺失 → 追加缺失列到表尾并保存；返回 (True, 追加后表头)
    - 文件损坏（BadZipFile 等）→ 备份损坏文件并重建

    Returns:
        (是否有改动, 追加后表头)
    """
    if not path.exists():
        ensure_exists(path)
        return True, list(HEADERS_ZH)

    try:
        current = _read_header(path)
    except (BadZipFile, Exception) as e:
        # manifest.xlsx 损坏（如 ZIP CRC-32 校验失败）→ 备份并重建
        log.error(
            "manifest 文件损坏，备份后重建: %s (%s)",
            path.name, e,
        )
        _backup_corrupted(path)
        ensure_exists(path)
        return True, list(HEADERS_ZH)

    if not current:
        # 空文件，重建
        ensure_exists(path)
        return True, list(HEADERS_ZH)

    missing = [h for h in HEADERS_ZH if h not in current]
    if not missing:
        return False, current

    # 追加缺失列
    wb = load_workbook(path)
    try:
        ws = wb.active
        next_col = len(current) + 1
        for h in missing:
            cell = ws.cell(row=1, column=next_col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EEEEEE")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(next_col)].width = max(
                12, len(h) * 2
            )
            # 给已有数据行填 None（占位）
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=next_col, value=None)
            next_col += 1
        wb.save(path)
        new_headers = current + missing
        return True, new_headers
    finally:
        wb.close()


def bootstrap(data_dir: Path) -> Path:
    """启动钩子：定位 manifest 文件，必要时补列。返回最终路径。"""
    path = find_manifest_file(data_dir)
    ensure_exists(path)
    ensure_columns(path)
    return path


# ============ 读 ============


def load(path: Path) -> Dict[str, ManifestRow]:
    """加载 manifest 为 {filename: ManifestRow} 字典。

    - 文件不存在 → 空字典
    - 表头按名字匹配，列顺序任意
    - 缺失列对应的字段为 None
    - 文件名 strip 规范化
    """
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {}
        header_list = [str(c).strip() if c is not None else "" for c in header_row]

        # 字段 → 列下标
        field_to_col: Dict[str, int] = {}
        for i, h in enumerate(header_list):
            f = HEADER_TO_FIELD.get(h)
            if f:
                field_to_col[f] = i

        if "filename" not in field_to_col:
            # 没有主键列
            return {}

        out: Dict[str, ManifestRow] = {}
        for raw in rows_iter:
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            row_dict = _raw_to_dict_by_field(field_to_col, raw)
            fname = row_dict.get("filename")
            if not fname:
                continue
            out[str(fname).strip()] = ManifestRow(**row_dict)
        return out
    finally:
        wb.close()


def _raw_to_dict_by_field(
    field_to_col: Dict[str, int], raw: Iterable
) -> Dict[str, object]:
    out: Dict[str, object] = {}
    for field, col_idx in field_to_col.items():
        if col_idx >= len(raw):
            out[field] = None
            continue
        value = raw[col_idx]
        if value is None:
            out[field] = None
        elif field == "seq":
            try:
                out[field] = int(value)
            except (TypeError, ValueError):
                out[field] = None
        else:
            # ★ 容错：用户 Excel 里偶尔把数字/日期等写错列（如把文件大小误填到"二级分类"），
            # 统一转 str 让 ManifestRow 验证通过。已 strip 去掉前后空白。
            if isinstance(value, str):
                out[field] = value.strip()
            else:
                out[field] = str(value).strip()
    return out


# ============ 写 ============


def _resolve_col_index(header_list: List[str], target_header: str) -> Optional[int]:
    """返回 target_header 在 header_list 中的 1-based 列号；找不到返回 None。"""
    for i, h in enumerate(header_list):
        if h == target_header:
            return i + 1
    return None


def upsert(path: Path, row: ManifestRow) -> None:
    """按 filename 插入或更新一行。

    - 调用前确保所有列已存在（upsert 内部会再 ensure_columns 一次以防新增列）
    - 按表头名匹配写入，列顺序无关
    - 跨进程并发由 openpyxl 不支持；进程内由 _write_lock 串行化
    """
    with _write_lock:
        ensure_exists(path)
        ensure_columns(path)
        wb = load_workbook(path)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_list = [str(c).strip() if c is not None else "" for c in header_row]

            # 查找已有行（按 filename 列）
            filename_col = _resolve_col_index(header_list, "文件名称")
            if filename_col is None:
                raise RuntimeError("manifest 缺少『文件名称』列")

            target_idx: Optional[int] = None
            for idx, existing in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if existing and len(existing) >= filename_col:
                    name_cell = existing[filename_col - 1]
                    if name_cell and str(name_cell).strip() == row.filename:
                        target_idx = idx
                        break

            # 写值：按 field → header → col 索引
            field_to_col = {}
            for i, h in enumerate(header_list):
                f = HEADER_TO_FIELD.get(h)
                if f:
                    field_to_col[f] = i + 1

            if target_idx is None:
                # 新行：按当前表头顺序写
                values = [None] * len(header_list)
                for field, col in field_to_col.items():
                    values[col - 1] = getattr(row, field, None)
                ws.append(values)
            else:
                for field, col in field_to_col.items():
                    value = getattr(row, field, None)
                    ws.cell(row=target_idx, column=col, value=value)
            wb.save(path)
        finally:
            wb.close()


def bulk_upsert(path: Path, rows: Iterable[ManifestRow]) -> None:
    """批量 upsert，合并到一次 workbook open/save。

    支持「文件名扩展名变化」的写入：
        - 优先按 row.filename 精确匹配
        - 否则按 stem 匹配（兼容 Excel 写「x」、磁盘是「x.pdf」的场景）
        - 都不匹配 → 追加新行
    """
    from pathlib import Path as _Path
    rows_list = list(rows)
    if not rows_list:
        return
    with _write_lock:
        ensure_exists(path)
        ensure_columns(path)
        wb = load_workbook(path)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_list = [str(c).strip() if c is not None else "" for c in header_row]

            filename_col = _resolve_col_index(header_list, "文件名称")
            if filename_col is None:
                raise RuntimeError("manifest 缺少『文件名称』列")

            # 已存在行索引：精确名 → 行号
            existing_idx: Dict[str, int] = {}
            # 同时记录 stem → 行号（处理扩展名变化）
            existing_stem_idx: Dict[str, int] = {}
            for idx, existing in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if existing and len(existing) >= filename_col:
                    name_cell = existing[filename_col - 1]
                    if name_cell:
                        name_str = str(name_cell).strip()
                        existing_idx[name_str] = idx
                        existing_stem_idx[_Path(name_str).stem] = idx

            # field → col
            field_to_col: Dict[str, int] = {}
            for i, h in enumerate(header_list):
                f = HEADER_TO_FIELD.get(h)
                if f:
                    field_to_col[f] = i + 1

            for row in rows_list:
                target_idx: Optional[int] = None
                if row.filename in existing_idx:
                    target_idx = existing_idx[row.filename]
                else:
                    row_stem = _Path(row.filename).stem
                    if row_stem in existing_stem_idx:
                        target_idx = existing_stem_idx[row_stem]
                        # 同步精确索引，便于后续行复用
                        existing_idx[row.filename] = target_idx
                        existing_stem_idx.pop(row_stem, None)

                if target_idx is not None:
                    for field, col in field_to_col.items():
                        value = getattr(row, field, None)
                        ws.cell(row=target_idx, column=col, value=value)
                else:
                    values = [None] * len(header_list)
                    for field, col in field_to_col.items():
                        values[col - 1] = getattr(row, field, None)
                    ws.append(values)
            wb.save(path)
        finally:
            wb.close()


# ============ 工具 ============


def now_iso() -> str:
    return _now_iso()
